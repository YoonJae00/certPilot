"""갭 리포트 XLSX 생성 (PRD §7 F7 일부).

시트 3개:

1. **요약** — 장별 met/partial/unmet/unknown 수와 준비도 %.
2. **항목별 판정** — 101행 전체.
3. **예상 결함** — 미충족·부분충족을 우선순위 순으로(unmet 먼저, 확신도 높은 순).

수치는 전부 `findings` 행에서 직접 집계한다. `summary_json` 을 그대로 베끼지 않는다
(리포트와 DB 집계가 어긋나는 걸 테스트로 잡을 수 있어야 한다).
"""

import io
import uuid
from dataclasses import dataclass

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet
from sqlalchemy import select
from sqlalchemy.orm import Session

from app.models import Assessment, Criterion, Finding, FindingStatus, Project
from app.services.scoring import code_sort_key, readiness_of

SHEET_SUMMARY = "요약"
SHEET_FINDINGS = "항목별 판정"
SHEET_DEFECTS = "예상 결함"

# 판정 표시 이름(사용자 문구는 한국어).
STATUS_LABELS: dict[FindingStatus, str] = {
    FindingStatus.MET: "충족",
    FindingStatus.PARTIAL: "부분충족",
    FindingStatus.UNMET: "미충족",
    FindingStatus.UNKNOWN: "판단불가",
}

DECIDED_BY_LABELS = {"rule": "규칙", "llm": "AI", "reviewer": "심사원"}

# 예상 결함 시트 정렬 우선순위.
_DEFECT_PRIORITY = {FindingStatus.UNMET: 0, FindingStatus.PARTIAL: 1}

_HEADER_FILL = PatternFill("solid", fgColor="1F3864")
_HEADER_FONT = Font(color="FFFFFF", bold=True)


@dataclass(frozen=True)
class ReportRow:
    """리포트 한 행에 필요한 값."""

    code: str
    chapter: int
    section: str
    title: str
    status: FindingStatus
    confidence: float
    decided_by: str
    predicted_defect: str
    recommendation: str
    rationale: str
    chunk_count: int
    evidence_count: int


def _load_rows(db: Session, assessment_id: uuid.UUID) -> list[ReportRow]:
    """판정과 인증기준을 조인해 리포트 행을 만든다."""
    records = db.execute(
        select(Finding, Criterion)
        .join(Criterion, Criterion.code == Finding.criterion_code)
        .where(Finding.assessment_id == assessment_id)
    ).all()

    rows = [
        ReportRow(
            code=finding.criterion_code,
            chapter=criterion.chapter,
            section=criterion.section,
            title=criterion.title,
            status=finding.status,
            confidence=round(float(finding.confidence), 2),
            decided_by=DECIDED_BY_LABELS.get(finding.decided_by.value, finding.decided_by.value),
            predicted_defect=finding.predicted_defect or "",
            recommendation=finding.recommendation or "",
            rationale=finding.rationale or "",
            chunk_count=len(finding.evidence_chunk_ids or []),
            evidence_count=len(finding.evidence_ids or []),
        )
        for finding, criterion in records
    ]
    rows.sort(key=lambda row: code_sort_key(row.code))
    return rows


def _write_header(sheet: Worksheet, headers: list[str], widths: list[int]) -> None:
    """머리글 행을 쓰고 열 너비를 맞춘다."""
    sheet.append(headers)
    for index, width in enumerate(widths, start=1):
        sheet.column_dimensions[get_column_letter(index)].width = width
    for cell in sheet[1]:
        cell.fill = _HEADER_FILL
        cell.font = _HEADER_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center")
    sheet.freeze_panes = "A2"


def _write_summary(
    sheet: Worksheet, rows: list[ReportRow], project: Project, assessment: Assessment
) -> None:
    """시트 1: 장별 집계와 준비도."""
    _write_header(
        sheet,
        ["장", "항목 수", "충족", "부분충족", "미충족", "판단불가", "준비도(%)"],
        [8, 10, 10, 12, 10, 12, 12],
    )

    chapters = sorted({row.chapter for row in rows})
    totals = {"total": 0, "met": 0, "partial": 0, "unmet": 0, "unknown": 0}

    for chapter in chapters:
        subset = [row for row in rows if row.chapter == chapter]
        counts = {status.value: 0 for status in FindingStatus}
        for row in subset:
            counts[row.status.value] += 1
        readiness = readiness_of(
            met=counts["met"],
            partial=counts["partial"],
            unknown=counts["unknown"],
            total=len(subset),
        )
        sheet.append(
            [
                f"{chapter}장",
                len(subset),
                counts["met"],
                counts["partial"],
                counts["unmet"],
                counts["unknown"],
                round(readiness * 100, 1),
            ]
        )
        totals["total"] += len(subset)
        for key in ("met", "partial", "unmet", "unknown"):
            totals[key] += counts[key]

    overall = readiness_of(
        met=totals["met"],
        partial=totals["partial"],
        unknown=totals["unknown"],
        total=totals["total"],
    )
    sheet.append(
        [
            "전체",
            totals["total"],
            totals["met"],
            totals["partial"],
            totals["unmet"],
            totals["unknown"],
            round(overall * 100, 1),
        ]
    )
    for cell in sheet[sheet.max_row]:
        cell.font = Font(bold=True)

    sheet.append([])
    sheet.append(["프로젝트", project.name])
    sheet.append(["인증 종류", project.cert_type.value])
    sheet.append(["모의심사 ID", str(assessment.id)])
    sheet.append(
        [
            "실행 완료",
            assessment.finished_at.strftime("%Y-%m-%d %H:%M") if assessment.finished_at else "",
        ]
    )
    sheet.append(["판정 모델", assessment.model or ""])
    sheet.append(
        [
            "준비도 산식",
            "(충족 + 0.5 × 부분충족) ÷ (항목 수 − 판단불가). 판단불가는 분모에서 뺀다.",
        ]
    )


def _write_findings(sheet: Worksheet, rows: list[ReportRow]) -> None:
    """시트 2: 항목별 판정 전체."""
    _write_header(
        sheet,
        [
            "항목 코드",
            "장",
            "분류",
            "항목명",
            "판정",
            "확신도",
            "판정 주체",
            "예상 결함",
            "개선 권고",
            "근거 청크",
            "근거 증적",
            "판정 근거",
        ],
        [12, 6, 26, 24, 10, 8, 10, 42, 42, 10, 10, 60],
    )
    for row in rows:
        sheet.append(
            [
                row.code,
                row.chapter,
                row.section,
                row.title,
                STATUS_LABELS[row.status],
                row.confidence,
                row.decided_by,
                row.predicted_defect,
                row.recommendation,
                row.chunk_count,
                row.evidence_count,
                row.rationale,
            ]
        )


def _write_defects(sheet: Worksheet, rows: list[ReportRow]) -> None:
    """시트 3: 예상 결함 우선순위(미충족 먼저, 확신도 높은 순)."""
    _write_header(
        sheet,
        ["우선순위", "항목 코드", "항목명", "판정", "확신도", "예상 결함", "개선 권고"],
        [10, 12, 24, 10, 8, 50, 50],
    )
    candidates = [row for row in rows if row.status in _DEFECT_PRIORITY]
    candidates.sort(
        key=lambda row: (_DEFECT_PRIORITY[row.status], -row.confidence, code_sort_key(row.code))
    )
    for priority, row in enumerate(candidates, start=1):
        sheet.append(
            [
                priority,
                row.code,
                row.title,
                STATUS_LABELS[row.status],
                row.confidence,
                row.predicted_defect,
                row.recommendation,
            ]
        )


def build_gap_report(db: Session, project: Project, assessment: Assessment) -> bytes:
    """갭 리포트 XLSX 를 만들어 바이트로 돌려준다."""
    rows = _load_rows(db, assessment.id)

    workbook = Workbook()
    summary_sheet = workbook.active
    if summary_sheet is None:  # pragma: no cover - openpyxl 은 항상 시트를 만든다
        summary_sheet = workbook.create_sheet()
    summary_sheet.title = SHEET_SUMMARY
    _write_summary(summary_sheet, rows, project, assessment)
    _write_findings(workbook.create_sheet(SHEET_FINDINGS), rows)
    _write_defects(workbook.create_sheet(SHEET_DEFECTS), rows)

    buffer = io.BytesIO()
    workbook.save(buffer)
    return buffer.getvalue()
