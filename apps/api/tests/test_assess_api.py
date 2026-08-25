"""모의심사 API·갭 리포트 테스트 (PRD §7 F3, F7 일부).

실제 LLM 은 절대 부르지 않는다. `.env` 에 키가 있어도 결정적 Fake 로 고정한다.
"""

import io
import time
import uuid

import pytest
from openpyxl import load_workbook
from sqlalchemy import select

from app.llm.provider import FakeProvider
from app.models import AuditLog, Finding, FindingStatus
from app.services.report import SHEET_DEFECTS, SHEET_FINDINGS, SHEET_SUMMARY
from app.workers.assess import run_assessment
from tests.conftest import login
from tests.test_assess import RULE_FAIL_CODE, create_assessment, seed_project

# 백그라운드 스레드 폴백이 끝나기를 기다리는 최대 시간(초).
POLL_TIMEOUT_SECONDS = 120


@pytest.fixture(autouse=True)
def _no_celery(monkeypatch):
    """업로드가 실제 브로커를 건드리지 않게 한다."""
    monkeypatch.setattr("app.api.documents.enqueue_ingest", lambda document_id: None)


@pytest.fixture(autouse=True)
def _fake_llm(monkeypatch):
    """비용이 드는 실제 LLM 호출을 원천 차단한다."""
    monkeypatch.setattr("app.workers.assess.get_llm_provider", FakeProvider)


@pytest.fixture
def assessed(client, db, tenants, storage):
    """완료된 모의심사 1건이 있는 A조직 프로젝트."""
    project_id = seed_project(client, db, tenants)
    assessment_id = create_assessment(db, project_id)
    run_assessment(assessment_id, provider=FakeProvider())
    return {"project_id": project_id, "assessment_id": assessment_id}


def wait_for_status(client, project_id, assessment_id, targets: set[str]) -> dict:
    """모의심사가 끝날 때까지 상세를 폴링한다."""
    deadline = time.monotonic() + POLL_TIMEOUT_SECONDS
    payload: dict = {}
    while time.monotonic() < deadline:
        response = client.get(f"/projects/{project_id}/assessments/{assessment_id}")
        assert response.status_code == 200, response.text
        payload = response.json()
        if payload["status"] in targets:
            return payload
        time.sleep(0.1)
    raise AssertionError(f"모의심사가 끝나지 않았다: 마지막 상태={payload.get('status')}")


def test_create_assessment_runs_and_completes(client, db, tenants, storage, monkeypatch):
    """202 로 큐잉되고, 브로커가 없으면 스레드 폴백으로 실행돼 done 이 된다."""
    project_id = seed_project(client, db, tenants)
    # 브로커가 없는 상황을 만든다(데모 환경).
    monkeypatch.setattr("app.api.assessments.enqueue_assessment", lambda assessment_id: False)

    response = client.post(f"/projects/{project_id}/assessments")
    assert response.status_code == 202
    created = response.json()
    assert created["status"] in {"queued", "running"}
    assessment_id = created["id"]

    payload = wait_for_status(client, project_id, assessment_id, {"done", "failed"})
    assert payload["status"] == "done"
    assert payload["model"] == FakeProvider.model_name
    assert payload["summary_json"]["progress"] == {"done": 101, "total": 101}
    assert sum(payload["summary_json"]["counts"].values()) == 101

    findings = client.get(f"/projects/{project_id}/assessments/{assessment_id}/findings")
    assert findings.status_code == 200
    assert len(findings.json()) == 101


def test_create_assessment_uses_celery_when_available(client, db, tenants, storage, monkeypatch):
    """브로커가 살아 있으면 큐에만 넣고 스레드 폴백을 쓰지 않는다."""
    project_id = seed_project(client, db, tenants)
    queued: list[uuid.UUID] = []
    monkeypatch.setattr(
        "app.api.assessments.enqueue_assessment",
        lambda assessment_id: queued.append(assessment_id) or True,
    )

    def _fail(assessment_id):  # pragma: no cover - 호출되면 테스트가 실패한다
        raise AssertionError("큐잉에 성공했는데 스레드 폴백이 돌았다")

    monkeypatch.setattr("app.api.assessments.start_assessment_thread", _fail)

    response = client.post(f"/projects/{project_id}/assessments")
    assert response.status_code == 202
    assert response.json()["status"] == "queued"
    assert queued == [uuid.UUID(response.json()["id"])]


def test_start_writes_audit_log(client, db, tenants, storage, monkeypatch):
    """모의심사 시작이 감사 로그에 남는다(사용자 포함)."""
    project_id = seed_project(client, db, tenants)
    monkeypatch.setattr("app.api.assessments.enqueue_assessment", lambda assessment_id: True)

    assessment_id = client.post(f"/projects/{project_id}/assessments").json()["id"]

    log = db.execute(
        select(AuditLog).where(
            AuditLog.action == "assessment.start", AuditLog.target == assessment_id
        )
    ).scalar_one()
    assert log.user_id == tenants["admin_a"].id
    assert log.org_id == tenants["org_a"].id


def test_list_assessments_is_newest_first(client, db, tenants, storage, monkeypatch):
    """목록은 최신순이다."""
    project_id = seed_project(client, db, tenants)
    monkeypatch.setattr("app.api.assessments.enqueue_assessment", lambda assessment_id: True)

    first = client.post(f"/projects/{project_id}/assessments").json()["id"]
    second = client.post(f"/projects/{project_id}/assessments").json()["id"]

    listed = client.get(f"/projects/{project_id}/assessments")
    assert listed.status_code == 200
    ids = [row["id"] for row in listed.json()]
    assert ids[:2] == [second, first]


def test_findings_filters_and_sort(client, assessed):
    """판정 목록 필터(status·chapter·q)와 정렬이 동작한다."""
    base = f"/projects/{assessed['project_id']}/assessments/{assessed['assessment_id']}/findings"

    everything = client.get(base).json()
    assert len(everything) == 101
    # 기본 정렬은 사람이 보는 항목 코드 순서다.
    assert everything[0]["criterion_code"] == "1.1.1"

    unknowns = client.get(base, params={"status": "unknown"}).json()
    assert unknowns
    assert {row["status"] for row in unknowns} == {"unknown"}

    chapter_three = client.get(base, params={"chapter": 3}).json()
    assert len(chapter_three) == 21
    assert {row["chapter"] for row in chapter_three} == {3}

    searched = client.get(base, params={"q": "사용자 인증"}).json()
    assert any(row["criterion_code"] == RULE_FAIL_CODE for row in searched)

    by_confidence = client.get(base, params={"sort": "-confidence"}).json()
    scores = [row["confidence"] for row in by_confidence]
    assert scores == sorted(scores, reverse=True)

    combined = client.get(base, params={"status": "unmet", "chapter": 2}).json()
    assert all(row["status"] == "unmet" and row["chapter"] == 2 for row in combined)


def test_findings_include_criterion_metadata(client, assessed):
    """판정 목록에 인증기준 제목·분류가 붙어 나온다."""
    base = f"/projects/{assessed['project_id']}/assessments/{assessed['assessment_id']}/findings"
    row = next(
        item
        for item in client.get(base).json()
        if item["criterion_code"] == RULE_FAIL_CODE
    )
    assert row["title"] == "사용자 인증"
    assert row["section"] == "2.5 인증 및 권한관리"
    assert row["status"] == "unmet"
    assert row["decided_by"] == "rule"


def test_finding_detail_contains_chunk_text_and_evidence(client, db, assessed):
    """상세에는 근거 청크 본문과 증적 payload 가 실린다(근거 하이라이트용)."""
    project_id = assessed["project_id"]
    assessment_id = assessed["assessment_id"]

    finding = db.execute(
        select(Finding).where(
            Finding.assessment_id == assessment_id,
            Finding.status != FindingStatus.UNKNOWN,
            Finding.criterion_code == RULE_FAIL_CODE,
        )
    ).scalar_one()

    response = client.get(
        f"/projects/{project_id}/assessments/{assessment_id}/findings/{finding.id}"
    )
    assert response.status_code == 200
    payload = response.json()

    assert payload["criterion_code"] == RULE_FAIL_CODE
    assert payload["criterion_requirement"]
    assert payload["evidence"], "규칙 fail 증적이 상세에 없다"
    assert payload["evidence"][0]["check_id"] == "mfa_enabled"
    assert payload["evidence"][0]["payload_json"]["users"] == 7

    if payload["evidence_chunk_ids"]:
        assert len(payload["chunks"]) == len(payload["evidence_chunk_ids"])
        for chunk in payload["chunks"]:
            assert chunk["text"]
            assert chunk["filename"]


def test_finding_detail_unknown_id_is_404(client, assessed):
    """없는 판정 id 는 404."""
    url = (
        f"/projects/{assessed['project_id']}/assessments/{assessed['assessment_id']}"
        f"/findings/{uuid.uuid4()}"
    )
    assert client.get(url).status_code == 404


def test_gap_report_xlsx_matches_db(client, db, assessed):
    """XLSX 리포트는 시트 3개이고 수치가 DB 집계와 일치한다."""
    project_id = assessed["project_id"]
    assessment_id = assessed["assessment_id"]

    response = client.get(
        f"/projects/{project_id}/assessments/{assessment_id}/report.xlsx"
    )
    assert response.status_code == 200
    assert response.headers["content-type"].startswith(
        "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    assert "attachment;" in response.headers["content-disposition"]

    workbook = load_workbook(io.BytesIO(response.content))
    assert workbook.sheetnames == [SHEET_SUMMARY, SHEET_FINDINGS, SHEET_DEFECTS]

    findings = list(
        db.execute(select(Finding).where(Finding.assessment_id == assessment_id)).scalars()
    )
    assert len(findings) == 101

    summary = workbook[SHEET_SUMMARY]
    rows = {row[0]: row for row in summary.iter_rows(min_row=2, max_row=5, values_only=True)}
    for chapter in (1, 2, 3):
        row = rows[f"{chapter}장"]
        expected = sum(
            1 for item in findings if item.criterion_code.startswith(f"{chapter}.")
        )
        assert row[1] == expected
    total_row = rows["전체"]
    assert total_row[1] == 101
    assert total_row[2] == sum(1 for item in findings if item.status is FindingStatus.MET)
    assert total_row[3] == sum(1 for item in findings if item.status is FindingStatus.PARTIAL)
    assert total_row[4] == sum(1 for item in findings if item.status is FindingStatus.UNMET)
    assert total_row[5] == sum(1 for item in findings if item.status is FindingStatus.UNKNOWN)

    detail = workbook[SHEET_FINDINGS]
    assert detail.max_row == 102  # 머리글 + 101행

    defects = workbook[SHEET_DEFECTS]
    gap_count = sum(
        1
        for item in findings
        if item.status in {FindingStatus.UNMET, FindingStatus.PARTIAL}
    )
    assert defects.max_row == gap_count + 1
    # 미충족이 부분충족보다 먼저 온다(우선순위).
    statuses = [row[3] for row in defects.iter_rows(min_row=2, values_only=True)]
    assert statuses == sorted(statuses, key=lambda label: 0 if label == "미충족" else 1)


def test_report_download_is_audited(client, db, assessed):
    """리포트 내려받기는 감사 로그에 남는다."""
    client.get(
        f"/projects/{assessed['project_id']}/assessments/{assessed['assessment_id']}/report.xlsx"
    )
    actions = list(
        db.execute(
            select(AuditLog.action).where(
                AuditLog.target == str(assessed["assessment_id"]),
                AuditLog.action == "assessment.report_download",
            )
        ).scalars()
    )
    assert actions


def test_other_org_cannot_see_assessment(client, assessed):
    """다른 조직은 모의심사·판정·리포트에 접근할 수 없다(404)."""
    project_id = assessed["project_id"]
    assessment_id = assessed["assessment_id"]

    login(client, "admin-b@example.com")
    assert client.get(f"/projects/{project_id}/assessments").status_code == 404
    assert client.get(f"/projects/{project_id}/assessments/{assessment_id}").status_code == 404
    assert (
        client.get(f"/projects/{project_id}/assessments/{assessment_id}/findings").status_code
        == 404
    )
    assert (
        client.get(
            f"/projects/{project_id}/assessments/{assessment_id}/report.xlsx"
        ).status_code
        == 404
    )
    assert client.post(f"/projects/{project_id}/assessments").status_code == 404


def test_reviewer_is_forbidden(client, assessed):
    """심사원은 조직 스코프 모의심사 API 를 쓸 수 없다(403)."""
    project_id = assessed["project_id"]
    assessment_id = assessed["assessment_id"]

    login(client, "reviewer@example.com")
    assert client.post(f"/projects/{project_id}/assessments").status_code == 403
    assert client.get(f"/projects/{project_id}/assessments").status_code == 403
    assert client.get(f"/projects/{project_id}/assessments/{assessment_id}").status_code == 403
    assert (
        client.get(f"/projects/{project_id}/assessments/{assessment_id}/findings").status_code
        == 403
    )


def test_member_can_read_but_not_start(client, assessed):
    """org_member 는 결과 열람만 되고 실행은 못 한다."""
    project_id = assessed["project_id"]
    assessment_id = assessed["assessment_id"]

    login(client, "member-a@example.com")
    assert client.post(f"/projects/{project_id}/assessments").status_code == 403
    assert client.get(f"/projects/{project_id}/assessments/{assessment_id}").status_code == 200
    assert (
        client.get(f"/projects/{project_id}/assessments/{assessment_id}/findings").status_code
        == 200
    )


def test_requires_login(client, assessed):
    """비로그인은 401."""
    project_id = assessed["project_id"]
    client.cookies.clear()
    assert client.get(f"/projects/{project_id}/assessments").status_code == 401
